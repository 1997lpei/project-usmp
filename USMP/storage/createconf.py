#!/bin/bash/env python
# -*- coding:utf-8 -*-
import sys
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter 

ldevlist_book_filepath = "uploadfiles\\storageconf\\"
ldevlist_book_filename = "ldevlist_book.xlsx"

#文件预处理
class preprocess():
    storage_ldev_dict = {}

    def __init__(self,var):
        self.args = var
        storage_ldev_dict = dict.fromkeys(self.args)

    def cut(self):
        for storage in self.args :
            filename = ldevlist_book_filepath + storage
            try:
                ldevlist_file = open(filename,"r+")
                ldevlist_info = ldevlist_file.read()
                ldevlist_file.close()
                if len(storage.split("_")[0]) == 5 :
                    str_split = "Serial#  : " + "3" +storage.split("_")[0]
                elif len(storage.split("_")[0]) == 6 :
                    str_split = "Serial#  : " + storage.split("_")[0]
                ldevlist_lines = ldevlist_info.split(str_split)
                ldev_list = []
                for ldevlist_line in ldevlist_lines :
                    if len(ldevlist_line) :
                        ldev_dict = {"LDEV":"","VOL_Capacity":"","PORTs":"","RAID_LEVEL":"","RAID_TYPE":"","RAID_GROUPs":""}
                        ldev_dict["LDEV"] = ldevlist_line.splitlines()[1].split(" : ")[1]
                        for ldevlist in ldevlist_line.splitlines() :
                            if "VOL_Capacity" in ldevlist:
                                ldev_dict["VOL_Capacity"] = ldevlist.split(" : ")[1]
                            elif "PORTs" in ldevlist :
                                ldev_dict["PORTs"] = ldevlist.split(" : ")[1:]
                            elif "RAID_LEVEL" in ldevlist :
                                ldev_dict["RAID_LEVEL"] = ldevlist.split(" : ")[1]
                            elif "RAID_TYPE" in ldevlist :
                                ldev_dict["RAID_TYPE"] = ldevlist.split(" : ")[1]
                            elif "RAID_GROUPs" in ldevlist :
                                ldev_dict["RAID_GROUPs"] = ldevlist.split(" : ")[1]
                            else :
                                continue
                        ldev_list.append(ldev_dict)
                self.storage_ldev_dict[storage] = ldev_list
            except Exception as e :
                print e
                raise e

    def createExcel(self):
        try:
            if (not (os.path.isfile(ldevlist_book_filepath+ldevlist_book_filename))) :
                ldevlist_book = Workbook()
                ldevlist_book.save(ldevlist_book_filepath+ldevlist_book_filename)
            ldevlist_book = load_workbook(ldevlist_book_filepath+ldevlist_book_filename)
            for storage in self.args :
                storage_name = str(storage.split("_")[0])
                if storage_name in ldevlist_book.sheetnames :
                    ldevlist_book.remove_sheet(ldevlist_book.get_sheet_by_name(storage_name))
                ldevlist_sheet = ldevlist_book.create_sheet(str(storage_name),0)
                ldevlist_sheet.append(["LDEV","VOL_Capacity(GB)","RAID_LEVEL","RAID_TYPE","RAID_GROUPs","PORTs","HOSTNAME","HLUN_ID"])
                for ldev_info in self.storage_ldev_dict[storage] :
                    LDEV = str(hex(int(ldev_info["LDEV"]))).upper().split("X")[1].zfill(4)   
                    VOL_Capacity = int(ldev_info["VOL_Capacity"])/1024/1024/2
                    RAID_LEVEL = ldev_info["RAID_LEVEL"]
                    RAID_TYPE = ldev_info["RAID_TYPE"]
                    RAID_GROUPs = ldev_info["RAID_GROUPs"]
                    if len(ldev_info["PORTs"]) :
                        for PORT in ldev_info["PORTs"] :
                            port_info = PORT.split(" ")
                            host = str(port_info[2])
                            hlun_id = str(port_info[1])
                            hport = str(port_info[0][0:5])
                            ldevlist_sheet.append([LDEV,VOL_Capacity,RAID_LEVEL,RAID_TYPE,RAID_GROUPs,hport,host,hlun_id])
                    else:
                        ldevlist_sheet.append([LDEV,VOL_Capacity,RAID_LEVEL,RAID_TYPE,RAID_GROUPs])
            ldevlist_book.save(ldevlist_book_filepath+ldevlist_book_filename)
        except Exception as e:
            print e
            raise e
        
        
        
    

if __name__ == '__main__' :
    storagelist_fil = open("storage.list","r+")
    storage_list = []
    for storage in storagelist_fil.readlines() :
        storage_list.append(storage.split("-->")[0])
    storagelist_fil.close()
    preprocess_storage = preprocess(storage_list).cut()
    preprocess_storage = preprocess(storage_list).createExcel()
    print "finished"